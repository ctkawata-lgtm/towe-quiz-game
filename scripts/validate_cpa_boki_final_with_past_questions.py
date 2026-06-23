from pathlib import Path

from openpyxl import load_workbook


OUTPUT = Path(
    r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database\財務会計論_簿記_cpa_過去問_最終_pdf_game_大問分割_過去問込み.xlsx"
)
PDF_TOTAL_PAGES = 975


def main():
    wb = load_workbook(OUTPUT, data_only=True)
    ws_q = wb["問題"]
    ws_p = wb["ページ情報"]

    problem_rows = list(ws_q.iter_rows(min_row=2, values_only=True))
    page_rows = {row[0]: row for row in ws_p.iter_rows(min_row=2, values_only=True)}

    missing_page_info = []
    invalid_pages = []
    too_many_subquestions = []
    tail_rows = []
    tail_bad_answers = []
    max_page = 0

    for row in problem_rows:
        problem_id = row[0]
        if not problem_id:
            continue

        page_row = page_rows.get(problem_id)
        if not page_row:
            missing_page_info.append(problem_id)
            continue

        values = [value for value in page_row[1:5] if isinstance(value, int)]
        if values:
            max_page = max(max_page, max(values))
        if any(value < 1 or value > PDF_TOTAL_PAGES for value in values):
            invalid_pages.append((problem_id, values))

        sub_count = 0
        answers = []
        for index in range(1, 26):
            base = 1 + (index - 1) * 2
            label = row[base]
            answer = row[base + 1]
            if label not in (None, ""):
                sub_count += 1
                answers.append(str(answer).strip() if answer is not None else "")
        if sub_count > 25:
            too_many_subquestions.append((problem_id, sub_count))

        if str(problem_id).startswith("CPA短答過去問"):
            tail_rows.append(problem_id)
            if len(answers) != 1 or answers[0] not in {"1", "2", "3", "4", "5", "6"}:
                tail_bad_answers.append((problem_id, answers))

    print(f"workbook={OUTPUT}")
    print(f"problem_rows={len(problem_rows)}")
    print(f"page_rows={len(page_rows)}")
    print(f"tail_rows={len(tail_rows)}")
    print(f"max_page={max_page}")
    print(f"missing_page_info={len(missing_page_info)}")
    print(f"invalid_pages={len(invalid_pages)}")
    print(f"too_many_subquestions={len(too_many_subquestions)}")
    print(f"tail_bad_answers={len(tail_bad_answers)}")
    print("last_tail_ids=" + ", ".join(tail_rows[-5:]))

    if missing_page_info or invalid_pages or too_many_subquestions or tail_bad_answers:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
