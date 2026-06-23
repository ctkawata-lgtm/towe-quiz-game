from pathlib import Path

import openpyxl


PATH = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database\財務会計論_簿記_cpa_過去問_最終_pdf_game.xlsx")
PDF_TOTAL_PAGES = 975


def main() -> None:
    wb = openpyxl.load_workbook(PATH, data_only=True)
    ws_q = wb["問題"]
    ws_p = wb["ページ情報"]

    ids = set()
    page_ids = set()
    bad_ranges = []
    too_many = []
    long_answers = []
    total_subanswers = 0

    for row in ws_q.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        problem_id = row[0]
        ids.add(problem_id)
        count = 0
        for index in range(1, 26):
            answer_col = index * 2
            answer = row[answer_col] if answer_col < len(row) else None
            if answer:
                count += 1
                total_subanswers += 1
                if len(str(answer)) > 16:
                    long_answers.append((problem_id, index, answer))
        if count > 25:
            too_many.append((problem_id, count))

    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        problem_id = row[0]
        page_ids.add(problem_id)
        q_start, q_end, a_start, a_end = row[2], row[3], row[4], row[5]
        if not (1 <= q_start <= q_end <= PDF_TOTAL_PAGES and 1 <= a_start <= a_end <= PDF_TOTAL_PAGES):
            bad_ranges.append((problem_id, q_start, q_end, a_start, a_end))
        if a_start <= q_end:
            bad_ranges.append((problem_id, q_start, q_end, a_start, a_end, "overlap"))

    print(f"sheets={wb.sheetnames}")
    print(f"question_rows={ws_q.max_row - 1}")
    print(f"page_rows={ws_p.max_row - 1}")
    print(f"subanswers={total_subanswers}")
    print(f"id_mismatch={len(ids ^ page_ids)}")
    print(f"bad_ranges={len(bad_ranges)}")
    print(f"too_many={len(too_many)}")
    print(f"long_answers={len(long_answers)}")
    print(f"confirm_rows={wb['要確認'].max_row - 1}")
    print(f"first_question={[cell.value for cell in ws_q[2]][:11]}")
    print(f"first_page_info={[cell.value for cell in ws_p[2]][:6]}")
    if ids ^ page_ids or bad_ranges or too_many or long_answers:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
