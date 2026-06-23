from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


INPUT = Path(
    r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database\財務会計論_簿記_cpa_過去問_最終_pdf_game_統合版.xlsx"
)
OUTPUT = Path(
    r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database\財務会計論_簿記_cpa_過去問_最終_pdf_game_統合版_ページ範囲修正版.xlsx"
)
PDF_TOTAL_PAGES = 975


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)


def main() -> None:
    wb = load_workbook(INPUT)
    ws = wb["ページ情報"]

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        problem_id = ws.cell(row_idx, 1).value
        if not problem_id:
            continue
        rows.append(
            {
                "row_idx": row_idx,
                "problem_id": str(problem_id),
                "q_start": ws.cell(row_idx, 3).value,
                "q_end": ws.cell(row_idx, 4).value,
                "a_start": ws.cell(row_idx, 5).value,
                "a_end": ws.cell(row_idx, 6).value,
            }
        )

    tail_rows = [row for row in rows if row["problem_id"].startswith("CPA短答過去問")]
    changes = []

    for index, row in enumerate(tail_rows):
        q_start = int(row["q_start"])
        q_end = int(row["q_end"])
        a_start = int(row["a_start"])
        a_end = int(row["a_end"])

        # In the tail past-question section, the question pages run until just
        # before the answer/explanation starts.
        fixed_q_end = max(q_end, a_start - 1)

        # The explanation pages run until just before the next question starts.
        # If multiple questions share the same next question page, keep the
        # current explicit end to avoid collapsing same-page records.
        fixed_a_end = a_end
        for later in tail_rows[index + 1 :]:
            next_q_start = int(later["q_start"])
            if next_q_start > q_start:
                fixed_a_end = max(a_end, next_q_start - 1)
                break

        fixed_q_end = min(max(fixed_q_end, q_start), PDF_TOTAL_PAGES)
        fixed_a_end = min(max(fixed_a_end, a_start), PDF_TOTAL_PAGES)

        if fixed_q_end != q_end:
            ws.cell(row["row_idx"], 4).value = fixed_q_end
            changes.append((row["problem_id"], "問題終了", q_end, fixed_q_end))
        if fixed_a_end != a_end:
            ws.cell(row["row_idx"], 6).value = fixed_a_end
            changes.append((row["problem_id"], "解説終了", a_end, fixed_a_end))

    # Add an audit sheet without disturbing the game-facing sheets.
    if "ページ範囲修正ログ" in wb.sheetnames:
        del wb["ページ範囲修正ログ"]
    log = wb.create_sheet("ページ範囲修正ログ")
    log.append(["問題番号", "項目", "修正前", "修正後"])
    for change in changes:
        log.append(list(change))
    for col in range(1, 5):
        clone_cell_style(ws.cell(1, min(col, ws.max_column)), log.cell(1, col))
    log.freeze_panes = "A2"
    log.auto_filter.ref = log.dimensions
    log.column_dimensions["A"].width = 28
    log.column_dimensions["B"].width = 14
    log.column_dimensions["C"].width = 12
    log.column_dimensions["D"].width = 12

    wb.save(OUTPUT)
    print(f"output={OUTPUT}")
    print(f"tail_rows={len(tail_rows)}")
    print(f"changes={len(changes)}")
    for item in changes[:30]:
        print(f"{item[0]} {item[1]} {item[2]} -> {item[3]}")
    if len(changes) > 30:
        print(f"... {len(changes) - 30} more")


if __name__ == "__main__":
    main()
