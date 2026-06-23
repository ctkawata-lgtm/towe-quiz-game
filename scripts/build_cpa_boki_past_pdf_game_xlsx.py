from __future__ import annotations

import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


sys.stdout.reconfigure(encoding="utf-8")
logging.getLogger("pypdf").setLevel(logging.ERROR)

PDF_PATH = Path(r"C:\Users\kawat\Downloads\26財務会計論（簿記）cpa、過去問.pdf")
OUT_DIR = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database")
OUT_PATH = OUT_DIR / "26財務会計論（簿記）cpa、過去問_pdf_game.xlsx"

MAX_SUBQUESTIONS = 20


@dataclass(frozen=True)
class SubAnswer:
    label: str
    answer: str


@dataclass(frozen=True)
class Stage:
    problem_id: str
    title: str
    question_start: int
    question_end: int
    answer_start: int
    answer_end: int
    subanswers: list[SubAnswer]


def is_auto_gradable_answer(answer: str) -> bool:
    value = answer.strip()
    if re.fullmatch(r"[△▲-]?\d[\d,]*(?:\.\d+)?", value):
        return True
    if re.fullmatch(r"[あ-んア-ンァ-ヶ一-龥A-Za-zＡ-Ｚａ-ｚ○×〇]{1,6}", value):
        return True
    return False


def compact(text: str) -> str:
    return re.sub(r"\s+", "", text or "")


def normalize_text(text: str) -> str:
    text = (text or "").replace("，", ",").replace("－", "-").replace("△ ", "△")
    text = re.sub(r"(?<=[△▲-])\s+(?=\d)", "", text)
    return re.sub(r"\s+", " ", text).strip()


def page_kind(text: str) -> str:
    head = compact(text[:700])
    if "解答解説" in head:
        return "answer"
    if "財務会計論" in head and "問題" in head:
        return "problem"
    return "other"


def internal_page_no(text: str) -> int | None:
    m = re.search(r"[―-]\s*(\d{1,3})\s*[―-]", text[:300])
    return int(m.group(1)) if m else None


def is_answer_key(text: str) -> bool:
    head = compact(text[:2500])
    return "＜解答＞" in head and "問題" in head


def title_from_answer_page(text: str, fallback_index: int) -> str:
    first = normalize_text(text[:500])
    m = re.search(r"(ＣＰＡ.+?財務会計論\(?計算\)?).*?解答解説", first)
    if m:
        return m.group(1).replace("ＣＰＡ", "CPA").strip(" -")
    m = re.search(r"(第[０-９0-9一二三四五六七八九十・]+回[^ ]*答練)", first)
    if m:
        return m.group(1)
    return f"CPA財務会計論（簿記）過去問 {fallback_index:03d}"


def find_problem_run(texts: list[str], start_page: int, key_page: int) -> tuple[int, int]:
    candidates: list[tuple[int, int, int | None]] = []
    for page_no in range(start_page, key_page):
        text = texts[page_no - 1]
        if page_kind(text) == "problem":
            candidates.append((page_no, page_no, internal_page_no(text)))
    if not candidates:
        return max(1, key_page - 1), key_page - 1

    runs: list[list[tuple[int, int, int | None]]] = []
    current: list[tuple[int, int, int | None]] = []
    prev_internal: int | None = None
    for item in candidates:
        page_no, _, current_internal = item
        restart = bool(
            current
            and current_internal is not None
            and prev_internal is not None
            and current_internal <= prev_internal
        )
        if restart:
            runs.append(current)
            current = []
        current.append(item)
        prev_internal = current_internal
    if current:
        runs.append(current)
    chosen = max(runs, key=lambda run: (len(run), run[-1][0]))
    return chosen[0][0], chosen[-1][0]


def find_answer_end(texts: list[str], key_page: int, next_key_page: int | None) -> int:
    limit = (next_key_page - 1) if next_key_page else len(texts)
    for page_no in range(key_page + 1, limit + 1):
        if page_kind(texts[page_no - 1]) == "problem":
            return page_no - 1
    return limit


def answer_key_section(text: str) -> str:
    normalized = normalize_text(text)
    start = re.search(r"＜\s*解\s*答\s*＞", normalized)
    if not start:
        return ""
    section = normalized[start.end() :]
    end = re.search(r"＜\s*解\s*説\s*＞|【\s*解説\s*】", section)
    if end:
        section = section[: end.start()]
    section = re.sub(r"問\s*題\s*\d+\s*[～~]\s*問\s*題\s*\d+\s*[:：][^問]+", " ", section)
    section = re.sub(r"各\s*\d+\s*点[×x]\s*\d+\s*箇所", " ", section)
    return section.strip()


def clean_answer(value: str) -> str:
    value = normalize_text(value)
    value = value.strip(" :：,，.。()（）")
    if re.fullmatch(r"[△▲-]?\d[\d,\s]*(?:\.\d+)?", value):
        value = re.sub(r"\s+", "", value)
    value = re.sub(r"(?<=\d),(?=\d{3}(?:\D|$))", ",", value)
    return value


def parse_answer_chunk(main_no: str, chunk: str) -> list[SubAnswer]:
    chunk = normalize_text(chunk)
    chunk = re.sub(r"^[:：\-\s]+", "", chunk)
    if not chunk:
        return []

    # Indexed answers: 問題 1  1 5,500  2 400,000 ...
    pair_matches = list(
        re.finditer(
            r"(?:^|\s)(\d{1,2})\s+([△▲-]?\d[\d,]*(?:\.\d+)?|[ぁ-んァ-ヶ一-龥A-Za-zＡ-Ｚａ-ｚ]+)",
            chunk,
        )
    )
    if pair_matches and pair_matches[0].group(1) == "1":
        answers: list[SubAnswer] = []
        for match in pair_matches:
            label = f"問題{main_no}-{match.group(1)}"
            answer = clean_answer(match.group(2))
            if answer:
                answers.append(SubAnswer(label, answer))
        if answers:
            return answers

    # Single answer: 問題 6 2,791
    m = re.search(r"([△▲-]?\d[\d,]*(?:\.\d+)?|[ぁ-んァ-ヶ一-龥A-Za-zＡ-Ｚａ-ｚ]+)", chunk)
    if not m:
        return []
    return [SubAnswer(f"問題{main_no}", clean_answer(m.group(1)))]


def parse_answer_key(text: str) -> list[SubAnswer]:
    section = answer_key_section(text)
    if not section:
        return []
    matches = list(re.finditer(r"問\s*題\s*(\d{1,2})", section))
    answers: list[SubAnswer] = []
    for index, match in enumerate(matches):
        main_no = match.group(1)
        end = matches[index + 1].start() if index + 1 < len(matches) else len(section)
        chunk = section[match.end() : end]
        answers.extend(parse_answer_chunk(main_no, chunk))
    return answers


def split_stage(base_id: str, title: str, q_start: int, q_end: int, a_start: int, a_end: int, answers: list[SubAnswer]) -> list[Stage]:
    stages: list[Stage] = []
    for offset in range(0, len(answers), MAX_SUBQUESTIONS):
        part = answers[offset : offset + MAX_SUBQUESTIONS]
        suffix = "" if len(answers) <= MAX_SUBQUESTIONS else f"_part{offset // MAX_SUBQUESTIONS + 1}"
        part_title = title if not suffix else f"{title}（小問{offset + 1}-{offset + len(part)}）"
        stages.append(Stage(f"{base_id}{suffix}", part_title, q_start, q_end, a_start, a_end, part))
    return stages


def collect_stages(texts: list[str]) -> tuple[list[Stage], list[str]]:
    key_pages = [idx for idx, text in enumerate(texts, start=1) if is_answer_key(text)]
    stages: list[Stage] = []
    skipped: list[str] = []
    search_start = 1
    for key_index, key_page in enumerate(key_pages, start=1):
        title = title_from_answer_page(texts[key_page - 1], key_index)
        answers = parse_answer_key(texts[key_page - 1])
        if not answers:
            skipped.append(f"key_page={key_page}: 解答一覧を抽出できませんでした: {title}")
            continue
        before_count = len(answers)
        rejected = [answer for answer in answers if not is_auto_gradable_answer(answer.answer)]
        answers = [answer for answer in answers if is_auto_gradable_answer(answer.answer)]
        for answer in rejected:
            skipped.append(f"key_page={key_page}: 自動採点しにくい解答を除外: {title} / {answer.label} = {answer.answer}")
        if not answers:
            skipped.append(f"key_page={key_page}: 抽出解答{before_count}件がすべて自動採点対象外でした: {title}")
            continue

        question_start, question_end = find_problem_run(texts, search_start, key_page)
        next_key_page = key_pages[key_index] if key_index < len(key_pages) else None
        answer_end = find_answer_end(texts, key_page, next_key_page)
        base_id = f"CPA簿記{key_index:03d}"
        stages.extend(split_stage(base_id, title, question_start, question_end, key_page, answer_end, answers))
        search_start = answer_end + 1
    return stages, skipped


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    side = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.border = Border(top=side, bottom=side, left=side, right=side)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_workbook(stages: list[Stage], skipped: list[str], total_pages: int) -> None:
    wb = openpyxl.Workbook()
    ws_q = wb.active
    ws_q.title = "問題"
    headers = ["問題番号"]
    for idx in range(1, MAX_SUBQUESTIONS + 1):
        headers.extend([f"{idx}問目の回答範囲", f"{idx}問目の正解"])
    ws_q.append(headers)
    for stage in stages:
        row = [stage.problem_id]
        for answer in stage.subanswers:
            row.extend([answer.label, answer.answer])
        row.extend([""] * (len(headers) - len(row)))
        ws_q.append(row)

    ws_p = wb.create_sheet("ページ情報")
    ws_p.append(
        [
            "問題番号",
            "タイトル",
            "PDF内問題開始ページ",
            "PDF内問題終了ページ",
            "PDF内解答開始ページ",
            "PDF内解答終了ページ",
        ]
    )
    for stage in stages:
        ws_p.append([stage.problem_id, stage.title, stage.question_start, stage.question_end, stage.answer_start, stage.answer_end])

    ws_h = wb.create_sheet("使い方")
    ws_h.append(["項目", "内容"])
    ws_h.append(["対象PDF", PDF_PATH.name])
    ws_h.append(["PDF総ページ", total_pages])
    ws_h.append(["登録ステージ数", len(stages)])
    ws_h.append(["抽出方針", "解答一覧ページから、数字・△付き数字・語句で入力できる小問を抽出しました。小問が20個を超える答練はpart分割しています。"])
    ws_h.append(["入力ルール", "カンマあり/なし、全角/半角はゲーム側でおおむね吸収されます。△の答えは△も入力してください。"])

    ws_s = wb.create_sheet("要確認")
    ws_s.append(["内容"])
    if skipped:
        for item in skipped:
            ws_s.append([item])
    else:
        ws_s.append(["解答一覧ページから抽出不能な答練は検出されませんでした。"])

    for ws in wb.worksheets:
        style_header(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(1, min(ws.max_row, 200) + 1))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 46)
    ws_q.column_dimensions["A"].width = 18
    ws_p.column_dimensions["A"].width = 18
    ws_p.column_dimensions["B"].width = 58
    ws_h.column_dimensions["B"].width = 90
    ws_s.column_dimensions["A"].width = 120

    wb.save(OUT_PATH)


def main() -> None:
    reader = PdfReader(str(PDF_PATH))
    texts = [page.extract_text() or "" for page in reader.pages]
    stages, skipped = collect_stages(texts)
    write_workbook(stages, skipped, len(texts))
    total_subanswers = sum(len(stage.subanswers) for stage in stages)
    multi_page_question = sum(1 for stage in stages if stage.question_end > stage.question_start)
    multi_page_answer = sum(1 for stage in stages if stage.answer_end > stage.answer_start)
    print(f"pdf_pages={len(texts)}")
    print(f"stages={len(stages)}")
    print(f"subanswers={total_subanswers}")
    print(f"skipped={len(skipped)}")
    print(f"multi_page_question_stages={multi_page_question}")
    print(f"multi_page_answer_stages={multi_page_answer}")
    print(f"output={OUT_PATH}")
    print("samples=")
    for stage in stages[:8]:
        print(
            f"  {stage.problem_id}: q={stage.question_start}-{stage.question_end} "
            f"a={stage.answer_start}-{stage.answer_end} sub={len(stage.subanswers)} "
            f"first={stage.subanswers[0].label}:{stage.subanswers[0].answer}"
        )
    if skipped:
        print("skipped_samples=")
        for item in skipped[:10]:
            print(f"  {item}")


if __name__ == "__main__":
    main()
