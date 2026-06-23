from pathlib import Path

import openpyxl


PATH = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database\財務会計論（計算）個別問題集_pdf_game.xlsx")
PDF_TOTAL_PAGES = 930


def main() -> None:
    wb = openpyxl.load_workbook(PATH, data_only=True)
    ws = wb["ページ情報"]

    bad_ranges = []
    overlaps = []
    duplicates = []
    seen = set()
    multi_problem = 0
    multi_answer = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        problem_id = row[0]
        problem_start, problem_end, answer_start, answer_end = row[2], row[3], row[4], row[5]
        if problem_id in seen:
            duplicates.append(problem_id)
        seen.add(problem_id)
        if not (
            1 <= problem_start <= problem_end <= PDF_TOTAL_PAGES
            and 1 <= answer_start <= answer_end <= PDF_TOTAL_PAGES
        ):
            bad_ranges.append((problem_id, problem_start, problem_end, answer_start, answer_end))
        if answer_start <= problem_end:
            overlaps.append((problem_id, problem_start, problem_end, answer_start, answer_end))
        if problem_end > problem_start:
            multi_problem += 1
        if answer_end > answer_start:
            multi_answer += 1

    print(f"sheets={wb.sheetnames}")
    print(f"question_rows={wb['問題'].max_row - 1}")
    print(f"page_rows={ws.max_row - 1}")
    print(f"duplicates={len(duplicates)}")
    print(f"bad_ranges={len(bad_ranges)}")
    print(f"overlaps={len(overlaps)}")
    print(f"multi_problem={multi_problem}")
    print(f"multi_answer={multi_answer}")
    print(f"first_row={[cell.value for cell in ws[2]][:10]}")
    print(f"last_row={[cell.value for cell in ws[ws.max_row]][:10]}")
    if duplicates or bad_ranges or overlaps:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
