from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from pypdf import PdfReader


BASE_DIR = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database")

TARGETS = [
    ("24_kakomonn_unei.pdf", "24_kakomonn_unei_pdf_game.xlsx"),
    ("24_kakomonn_chusyou.pdf", "24_kakomonn_chusyou_pdf_game.xlsx"),
    ("24_kakomonn_jouhou.pdf", "24_kakomonn_jouhou_pdf_game.xlsx"),
    ("24_kakomonn_keizai.pdf", "24_kakomonn_keizai_pdf_game.xlsx"),
]


def norm(value) -> str:
    return str(value or "").strip().replace(" ", "").replace("\u3000", "")


def col_by_header(ws, names: list[str]) -> int | None:
    normalized = {norm(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    for name in names:
        col = normalized.get(norm(name))
        if col:
            return col
    return None


def ensure_col(ws, name: str) -> int:
    col = col_by_header(ws, [name])
    if col:
        return col
    col = ws.max_column + 1
    ws.cell(1, col).value = name
    return col


def question_number(question_id: str) -> int:
    match = re.search(r"-(\d+)$", str(question_id))
    if not match:
        raise ValueError(f"問題番号から問番号を読めません: {question_id}")
    return int(match.group(1))


def build_heading_finder(reader: PdfReader):
    cache: dict[tuple[int, int], float | None] = {}

    def heading_y(page_no: int, no: int) -> float | None:
        key = (page_no, no)
        if key in cache:
            return cache[key]
        hits: list[float] = []

        def visitor(text, cm, tm, font_dict, font_size):
            for match in re.finditer(r"第\s*(\d+)\s*問", text):
                if int(match.group(1)) == no:
                    hits.append(float(tm[5]))

        reader.pages[page_no - 1].extract_text(visitor_text=visitor)
        cache[key] = max(hits) if hits else None
        return cache[key]

    return heading_y


def visible_text(reader: PdfReader, page_no: int) -> str:
    return (reader.pages[page_no - 1].extract_text() or "").replace("\x00", "").strip()


def fix_workbook(pdf_name: str, xlsx_name: str) -> dict:
    pdf_path = BASE_DIR / pdf_name
    xlsx_path = BASE_DIR / xlsx_name
    output_path = xlsx_path.with_name(f"{xlsx_path.stem}_pagefix.xlsx")

    reader = PdfReader(str(pdf_path))
    max_page = len(reader.pages)
    heading_y = build_heading_finder(reader)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["ページ情報"]

    id_col = col_by_header(ws, ["問題番号"])
    qs_col = col_by_header(ws, ["PDF内問題開始ページ", "問題開始ページ"])
    qe_col = col_by_header(ws, ["PDF内問題終了ページ", "問題終了ページ"])
    as_col = col_by_header(ws, ["PDF内解答開始ページ", "解答ページ", "答えページ"])
    ae_col = col_by_header(ws, ["PDF内解答終了ページ", "解答終了ページ"])
    if not id_col or not qs_col or not qe_col or not as_col:
        raise ValueError(f"{xlsx_name}: ページ情報シートの必須列が不足しています。")
    if not ae_col:
        ae_col = ensure_col(ws, "解答終了ページ")
        for row in range(2, ws.max_row + 1):
            ws.cell(row, ae_col).value = ws.cell(row, as_col).value

    rows: list[dict] = []
    for row in range(2, ws.max_row + 1):
        qid = str(ws.cell(row, id_col).value).strip()
        if not qid:
            continue
        rows.append({
            "row": row,
            "id": qid,
            "no": question_number(qid),
            "qs": int(ws.cell(row, qs_col).value),
            "qe": int(ws.cell(row, qe_col).value),
            "as": int(ws.cell(row, as_col).value),
            "ae": int(ws.cell(row, ae_col).value),
        })

    changes: list[tuple[str, str, int, int]] = []
    for kind, start_key, end_key, threshold, end_col in [
        ("問題", "qs", "qe", 720, qe_col),
        ("解答", "as", "ae", 760, ae_col),
    ]:
        for index, row in enumerate(rows):
            start_page = row[start_key]
            old_end = row[end_key]
            new_end = start_page
            if index + 1 < len(rows):
                next_row = rows[index + 1]
                next_start = next_row[start_key]
                if next_start == start_page:
                    new_end = start_page
                elif next_start > start_page:
                    next_y = heading_y(next_start, next_row["no"])
                    new_end = next_start if (next_y is not None and next_y < threshold) else next_start - 1
            elif kind == "解答":
                page = start_page
                while page + 1 <= max_page and visible_text(reader, page + 1):
                    page += 1
                new_end = page
            if new_end != old_end:
                ws.cell(row["row"], end_col).value = new_end
                changes.append((kind, row["id"], old_end, new_end))

    wb.save(output_path)
    return {
        "pdf": pdf_name,
        "xlsx": xlsx_name,
        "output": str(output_path),
        "rows": len(rows),
        "max_page": max_page,
        "changes": changes,
    }


def main() -> None:
    for pdf_name, xlsx_name in TARGETS:
        result = fix_workbook(pdf_name, xlsx_name)
        print(result["output"])
        print(f"rows={result['rows']} pdf_pages={result['max_page']} changes={len(result['changes'])}")
        for change in result["changes"]:
            print("  ", change)


if __name__ == "__main__":
    main()
