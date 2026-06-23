from __future__ import annotations

import copy
import re
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas


SOURCE_PDF = Path(r"C:\AIroom\codexcli\財務諸表理論アプリ\assets\pdfs\財務諸表論問題集.pdf")
OUT_DIR = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database")
OUTPUT_PDF = OUT_DIR / "財務諸表論問題集_page_game.pdf"
OUTPUT_XLSX = OUT_DIR / "財務諸表論問題集_page_game.xlsx"

PROBLEM_CROP_LEFT = 39.7
PROBLEM_CROP_RIGHT = 260.0
LATE_PROBLEM_CROP_RIGHT = 275.0
LATE_PROBLEM_START_PAGE = 31
WIDE_PROBLEM_CROP_RIGHT = 290.0
WIDE_PROBLEM_START_PAGE = 35
WIDE_PROBLEM_EXCEPTIONS = {38, 39, 45}
NARROW_PROBLEM_CROP_RIGHT = 270.0
NARROW_PROBLEM_EXCEPTIONS = {46, 47}
ANSWER_CROP_RIGHT_PADDING = 8
PROBLEM_CROP_BOTTOM = 113.4
QUESTION_MASK_X = 250
TABLE_TOP = 790
TABLE_BOTTOM = 105

ANSWER_OVERRIDES = {
    60: {1: "1", 2: "0", 3: "0", 4: "0", 5: "0", 6: "1", 7: "0", 8: "1", 9: "1", 10: "1"},
    76: {1: "0", 2: "0", 3: "0", 4: "0", 5: "0", 6: "1", 7: "1", 8: "0", 9: "1"},
    78: {1: "0", 2: "0", 3: "1", 4: "1", 5: "0", 6: "1"},
}


def normalize_text(text: str) -> str:
    text = text.replace("\x00", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def extract_table_pages(reader: PdfReader) -> list[dict]:
    pages: list[dict] = []
    current_section = ""
    for page_no, page in enumerate(reader.pages, start=1):
        text = normalize_text((page.extract_text() or "").replace("\n", " "))
        if "番号 問題 解答欄 解説" not in text:
            continue
        title_match = re.search(r"第\s*([0-9０-９]+)\s*回\s*([^［\[]+)", text)
        if title_match:
            current_section = title_match.group(0).strip()

        numbers: list[tuple[int, float]] = []
        answer_marks: list[tuple[str, float]] = []

        def visitor(fragment, cm, tm, font_dict, font_size):
            value = fragment.strip()
            x = float(tm[4])
            y = float(tm[5])
            if value.isdigit() and 70 <= x <= 95 and 80 <= y <= 780:
                numbers.append((int(value), y))
            elif value in {"○", "〇", "×"} and 100 <= x <= 380 and 80 <= y <= 780:
                answer_marks.append((value, y))

        page.extract_text(visitor_text=visitor)
        number_positions = sorted(set(numbers), key=lambda item: -item[1])
        numbers_in_page = [number for number, _y in number_positions]
        rows = extract_page_answers(text, numbers_in_page)
        rows = merge_coordinate_answers(rows, number_positions, answer_marks)
        if page_no in ANSWER_OVERRIDES:
            rows = [
                {"number": number, "answer_symbol": "○" if value == "1" else "×", "answer_value": value}
                for number, value in ANSWER_OVERRIDES[page_no].items()
            ]
        if rows:
            pages.append({"source_page": page_no, "title": current_section or f"元PDF {page_no}", "rows": rows})
    return pages


def extract_page_answers(text: str, numbers: list[int]) -> list[dict]:
    rows: list[dict] = []
    if not numbers:
        return rows
    pattern = re.compile(r"(?:^|\s)(\d{1,2})\s+(.+?)(?=\s+\d{1,2}\s+|$)")
    chunks = {int(match.group(1)): match.group(2).strip() for match in pattern.finditer(text) if int(match.group(1)) in numbers}
    for number in numbers:
        chunk = chunks.get(number, "")
        answer_match = re.search(r"[\u25cb\u3007\u00d7]", chunk)
        if not answer_match:
            continue
        symbol = answer_match.group(0)
        rows.append({"number": number, "answer_symbol": "○" if symbol in {"○", "〇"} else "×", "answer_value": "1" if symbol in {"○", "〇"} else "0"})
    return rows


def merge_coordinate_answers(rows: list[dict], number_positions: list[tuple[int, float]], answer_marks: list[tuple[str, float]]) -> list[dict]:
    by_number = {row["number"]: row for row in rows}
    used: set[int] = set()
    for number, y in number_positions:
        if number in by_number:
            continue
        candidates = [
            (idx, symbol, abs(y - answer_y))
            for idx, (symbol, answer_y) in enumerate(answer_marks)
            if idx not in used and abs(y - answer_y) <= 15
        ]
        if not candidates:
            continue
        idx, symbol, _distance = min(candidates, key=lambda item: item[2])
        used.add(idx)
        by_number[number] = {"number": number, "answer_symbol": "○" if symbol in {"○", "〇"} else "×", "answer_value": "1" if symbol in {"○", "〇"} else "0"}
    return [by_number[number] for number, _y in number_positions if number in by_number]


def white_overlay(width: float, height: float, rectangles: list[tuple[float, float, float, float]]):
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=(width, height))
    c.setFillColorRGB(1, 1, 1)
    c.setStrokeColorRGB(1, 1, 1)
    for x, y, w, h in rectangles:
        c.rect(x, y, w, h, fill=1, stroke=0)
    c.save()
    packet.seek(0)
    return PdfReader(packet).pages[0]


def masked_problem_page(reader: PdfReader, page_no: int, output_problem_index: int):
    page = copy.copy(reader.pages[page_no - 1])
    width = float(page.mediabox.width)
    height = float(page.mediabox.height)
    if output_problem_index in NARROW_PROBLEM_EXCEPTIONS:
        crop_right = NARROW_PROBLEM_CROP_RIGHT
        mask_x = NARROW_PROBLEM_CROP_RIGHT
    elif output_problem_index in WIDE_PROBLEM_EXCEPTIONS:
        crop_right = LATE_PROBLEM_CROP_RIGHT
        mask_x = LATE_PROBLEM_CROP_RIGHT
    elif output_problem_index >= WIDE_PROBLEM_START_PAGE:
        crop_right = WIDE_PROBLEM_CROP_RIGHT
        mask_x = WIDE_PROBLEM_CROP_RIGHT
    elif output_problem_index >= LATE_PROBLEM_START_PAGE:
        crop_right = LATE_PROBLEM_CROP_RIGHT
        mask_x = LATE_PROBLEM_CROP_RIGHT
    else:
        crop_right = PROBLEM_CROP_RIGHT
        mask_x = QUESTION_MASK_X
    page.cropbox.lower_left = (PROBLEM_CROP_LEFT, PROBLEM_CROP_BOTTOM)
    page.cropbox.upper_right = (crop_right, height)
    overlay = white_overlay(width, height, [(mask_x, TABLE_BOTTOM, width - mask_x - 25, TABLE_TOP - TABLE_BOTTOM)])
    page.merge_page(overlay)
    return page


def full_answer_page(reader: PdfReader, page_no: int):
    page = copy.copy(reader.pages[page_no - 1])
    full_box = copy.copy(page.mediabox)
    page.cropbox = full_box
    page.trimbox = copy.copy(full_box)
    page.bleedbox = copy.copy(full_box)
    page.artbox = copy.copy(full_box)
    return page


def build_pdf(reader: PdfReader, pages: list[dict]) -> None:
    writer = PdfWriter()
    writer.add_metadata({"/Title": "財務諸表論問題集 page game"})
    for index, page_info in enumerate(pages, start=1):
        writer.add_page(masked_problem_page(reader, page_info["source_page"], index))
        writer.add_outline_item(f"問題 {index:02d} {page_info['title']}", len(writer.pages) - 1)
    for index, page_info in enumerate(pages, start=1):
        writer.add_page(full_answer_page(reader, page_info["source_page"]))
        writer.add_outline_item(f"解答 {index:02d} {page_info['title']}", len(writer.pages) - 1)
    with OUTPUT_PDF.open("wb") as file:
        writer.write(file)


def build_excel(pages: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws_answers = wb.active
    ws_answers.title = "問題"
    ws_pages = wb.create_sheet("ページ情報")
    ws_usage = wb.create_sheet("使い方")

    answer_headers = ["問題番号"]
    for i in range(1, 21):
        answer_headers += [f"{i}問目の回答範囲", f"{i}問目の正解"]
    page_headers = [
        "問題番号",
        "PDF内問題開始ページ",
        "PDF内問題終了ページ",
        "PDF内解答開始ページ",
        "PDF内解答終了ページ",
        "書籍印刷問題開始ページ",
        "書籍印刷問題終了ページ",
        "書籍印刷解答開始ページ",
        "書籍印刷解答終了ページ",
    ]
    ws_answers.append(answer_headers)
    ws_pages.append(page_headers)

    offset = len(pages)
    for index, page_info in enumerate(pages, start=1):
        qid = f"財表P{index:03d}"
        row = [qid] + [None] * 40
        for sub_index, item in enumerate(page_info["rows"], start=1):
            row[1 + (sub_index - 1) * 2] = f"第{item['number']}問"
            if item["answer_value"] == "1":
                row[2 + (sub_index - 1) * 2] = "1|○|まる|o"
            else:
                row[2 + (sub_index - 1) * 2] = "0|×|ばつ|x"
        ws_answers.append(row)
        answer_page = offset + index
        ws_pages.append([qid, index, index, answer_page, answer_page, index, index, answer_page, answer_page])

    ws_usage.append(["財務諸表論問題集 ページ単位ゲーム用データ", "PDF + Excel モード"])
    ws_usage.append(["対象PDF", OUTPUT_PDF.name])
    ws_usage.append(["問題単位", "PDFの1ページを1問として扱い、ページ内の○×問題を小問として登録。"])
    ws_usage.append(["入力ルール", "1=○、0=×。別解として○/×/まる/ばつ/o/xも登録。"])
    ws_usage.append(["PDF構成", f"前半{len(pages)}ページが問題、後半{len(pages)}ページが解答解説。"])
    style_workbook(wb)
    wb.save(OUTPUT_XLSX)


def style_workbook(wb: openpyxl.Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            width = 14
            if col == 1:
                width = 14
            elif ws.title == "問題":
                width = 16 if col % 2 == 0 else 18
            elif ws.title == "ページ情報":
                width = 20
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.auto_filter.ref = ws.dimensions


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    reader = PdfReader(str(SOURCE_PDF))
    pages = extract_table_pages(reader)
    if not pages:
        raise RuntimeError("No table pages detected.")
    overflow = [page for page in pages if len(page["rows"]) > 20]
    if overflow:
        raise RuntimeError(f"Pages with over 20 rows: {[(p['source_page'], len(p['rows'])) for p in overflow]}")
    build_pdf(reader, pages)
    build_excel(pages)
    print(OUTPUT_PDF)
    print(OUTPUT_XLSX)
    print(f"page_questions={len(pages)}")
    print(f"sub_questions={sum(len(p['rows']) for p in pages)}")
    print("first", pages[0])
    print("last", pages[-1])


if __name__ == "__main__":
    main()
