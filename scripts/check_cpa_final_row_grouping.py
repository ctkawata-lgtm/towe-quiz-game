from pathlib import Path
import re

import openpyxl


PATH = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database\財務会計論_簿記_cpa_過去問_最終_pdf_game.xlsx")
DIGIT_TRANS = str.maketrans("０１２３４５６７８９", "0123456789")


def main_problem(label: object) -> str | None:
    match = re.search(r"問題\s*([0-9０-９]+)", str(label or ""))
    if not match:
        return None
    return str(int(match.group(1).translate(DIGIT_TRANS)))


def main() -> None:
    wb = openpyxl.load_workbook(PATH, data_only=True)
    ws = wb["問題"]
    mixed = []
    samples = []
    rows_with_answers = 0

    for row_no in range(2, ws.max_row + 1):
        ranges = [
            ws.cell(row_no, col_no).value
            for col_no in range(2, ws.max_column + 1, 2)
            if ws.cell(row_no, col_no).value
        ]
        problem_numbers = [num for num in (main_problem(value) for value in ranges) if num]
        if problem_numbers:
            rows_with_answers += 1
        unique = sorted(set(problem_numbers), key=int)
        if len(unique) > 1:
            mixed.append((row_no, ws.cell(row_no, 1).value, ranges[:12], unique))
        if row_no <= 9:
            samples.append((row_no, ws.cell(row_no, 1).value, ranges[:12], unique))

    print(f"rows_with_answers={rows_with_answers}")
    print(f"mixed_count={len(mixed)}")
    print("samples=")
    for item in samples:
        print(item)
    print("first_mixed=")
    for item in mixed[:10]:
        print(item)
    if mixed:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
