from __future__ import annotations

import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

import build_cpa_boki_final_pdf_game_xlsx as regular
import build_cpa_boki_past_pdf_game_xlsx as base


sys.stdout.reconfigure(encoding="utf-8")
logging.getLogger("pypdf").setLevel(logging.ERROR)

PDF_PATH = Path(r"C:\Users\kawat\Downloads\財務会計論_簿記_cpa_過去問_最終.pdf")
OUT_PATH = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database\財務会計論_簿記_cpa_過去問_最終_pdf_game_大問分割_過去問込み.xlsx")
TAIL_START_PAGE = 720


@dataclass(frozen=True)
class TailQuestion:
    problem_id: str
    title: str
    question_page: int
    answer_page: int
    answer: str


def compact(text: str) -> str:
    return re.sub(r"\s+", "", text or "").translate(str.maketrans("０１２３４５６７８９－", "0123456789-"))


def visible(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def question_markers(text: str) -> list[tuple[str, str]]:
    normalized = visible(text)
    markers: list[tuple[str, str]] = []
    for match in re.finditer(r"(?:【([^】]{0,40}?問題\s*\d{1,2})】|(?:^|\s)(問題\s*\d{1,2})(?=\s+重要度))", normalized):
        label = match.group(1) or match.group(2)
        if not label:
            continue
        number = re.search(r"問題\s*(\d{1,2})", label)
        if not number:
            continue
        markers.append((str(int(number.group(1))), label.strip()))
    return markers


def answer_markers(text: str) -> list[str]:
    c = compact(text)
    values = []
    for match in re.finditer(r"正解([1-6１-６])", c):
        values.append(match.group(1).translate(str.maketrans("１２３４５６", "123456")))
    return values


def build_tail_questions(texts: list[str]) -> tuple[list[TailQuestion], list[str]]:
    question_pages: list[tuple[int, str, str]] = []
    for page_no in range(TAIL_START_PAGE, len(texts) + 1):
        text = texts[page_no - 1]
        if "正解" in compact(text):
            continue
        for number, title in question_markers(text):
            question_pages.append((page_no, number, title))

    answer_pages: list[tuple[int, str]] = []
    for page_no in range(TAIL_START_PAGE, len(texts) + 1):
        for answer in answer_markers(texts[page_no - 1]):
            answer_pages.append((page_no, answer))

    rows: list[TailQuestion] = []
    skipped: list[str] = []
    if len(question_pages) != len(answer_pages):
        skipped.append(f"tail_count_mismatch: 問題マーカー{len(question_pages)}件 / 正解マーカー{len(answer_pages)}件")

    for index, question in enumerate(question_pages, start=1):
        question_page, number, title = question
        if index > len(answer_pages):
            skipped.append(f"tail_question_page={question_page}: 正解ページが見つかりませんでした: {title}")
            continue
        answer_page, answer_value = answer_pages[index - 1]
        if answer_page <= question_page:
            skipped.append(f"tail_pair_order_warning: 問題{question_page}より前/同ページに正解{answer_page}: {title}")
        rows.append(
            TailQuestion(
                problem_id=f"CPA短答過去問{index:03d}_Q{int(number):02d}",
                title=title,
                question_page=question_page,
                answer_page=answer_page,
                answer=answer_value,
            )
        )
    return rows, skipped


def write_combined_workbook(stages: list[base.Stage], skipped: list[str], tail_rows: list[TailQuestion], tail_skipped: list[str], total_pages: int) -> None:
    wb = openpyxl.Workbook()
    ws_q = wb.active
    ws_q.title = "問題"
    headers = ["問題番号"]
    for idx in range(1, 26):
        headers.extend([f"{idx}問目の回答範囲", f"{idx}問目の正解"])
    ws_q.append(headers)

    for stage in stages:
        row = [stage.problem_id]
        for answer in stage.subanswers:
            row.extend([answer.label, answer.answer])
        row.extend([""] * (len(headers) - len(row)))
        ws_q.append(row)

    for row_item in tail_rows:
        row = [row_item.problem_id, row_item.title, row_item.answer]
        row.extend([""] * (len(headers) - len(row)))
        ws_q.append(row)

    ws_p = wb.create_sheet("ページ情報")
    ws_p.append(["問題番号", "タイトル", "PDF内問題開始ページ", "PDF内問題終了ページ", "PDF内解答開始ページ", "PDF内解答終了ページ"])
    for stage in stages:
        ws_p.append([stage.problem_id, stage.title, stage.question_start, stage.question_end, stage.answer_start, stage.answer_end])
    for row_item in tail_rows:
        ws_p.append([row_item.problem_id, row_item.title, row_item.question_page, row_item.question_page, row_item.answer_page, row_item.answer_page])

    ws_h = wb.create_sheet("使い方")
    ws_h.append(["項目", "内容"])
    ws_h.append(["対象PDF", PDF_PATH.name])
    ws_h.append(["PDF総ページ", total_pages])
    ws_h.append(["登録ステージ数", len(stages) + len(tail_rows)])
    ws_h.append(["登録小問数", sum(len(stage.subanswers) for stage in stages) + len(tail_rows)])
    ws_h.append(["抽出方針", "前半の答練は大問単位、後半の短答過去問は1問1ステージとして登録。短答過去問は選択肢1〜6の正解番号を解説ページの「正解」から取得。"])

    ws_s = wb.create_sheet("要確認")
    ws_s.append(["内容"])
    for item in skipped + tail_skipped:
        ws_s.append([item])
    if ws_s.max_row == 1:
        ws_s.append(["抽出エラーは検出されませんでした。"])

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    side = Side(style="thin", color="D9E2F3")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.border = Border(top=side, bottom=side, left=side, right=side)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(1, min(ws.max_row, 200) + 1))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 58)
    ws_p.column_dimensions["B"].width = 64
    ws_s.column_dimensions["A"].width = 120

    wb.save(OUT_PATH)


def main() -> None:
    reader = PdfReader(str(PDF_PATH))
    texts = [page.extract_text() or "" for page in reader.pages]
    stages, skipped = regular.collect_stages(texts)
    tail_rows, tail_skipped = build_tail_questions(texts)
    write_combined_workbook(stages, skipped, tail_rows, tail_skipped, len(texts))
    print(f"regular_stages={len(stages)}")
    print(f"tail_questions={len(tail_rows)}")
    print(f"tail_skipped={len(tail_skipped)}")
    print(f"total_rows={len(stages) + len(tail_rows)}")
    print(f"output={OUT_PATH}")
    print("tail_samples=")
    for row in tail_rows[:10]:
        print(f"  {row.problem_id}: q={row.question_page} a={row.answer_page} ans={row.answer} title={row.title}")


if __name__ == "__main__":
    main()
