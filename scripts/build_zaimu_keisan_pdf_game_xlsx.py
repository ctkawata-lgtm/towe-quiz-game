from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database")
PDF_PATH = BASE / "財務会計論（計算）個別問題集.pdf"
OUT_PATH = BASE / "財務会計論（計算）個別問題集_pdf_game.xlsx"

LABEL_RE = re.compile(r"[－-]\s*(\d+)\s*-\s*(\d+)\s*[－-]")
TITLE_RE = re.compile(r"[－-]\s*\d+\s*-\s*\d+\s*[－-]\s*[－-]\s*(.+?)\s*[－-]")
STANDALONE_TITLE_RE = re.compile(r"[－-]\s*([^－\n]*?（重要度：[^）]+難易度：[^）]+）)\s*[－-]")
ANSWER_RE = re.compile(r"解\s*答")


@dataclass(frozen=True)
class PageRecord:
    pdf_page: int
    chapter: int
    seq: int
    titles: tuple[str, ...]
    is_answer_start: bool

    @property
    def printed_label(self) -> str:
        return f"{self.chapter}-{self.seq}"

    @property
    def title(self) -> str:
        return self.titles[0] if self.titles else ""


@dataclass(frozen=True)
class StartRecord:
    pdf_page: int
    printed_label: str
    title: str


@dataclass(frozen=True)
class GameRow:
    problem_id: str
    title: str
    problem_start_page: int
    problem_end_page: int
    answer_start_page: int
    answer_end_page: int
    problem_start_label: str
    problem_end_label: str
    answer_start_label: str
    answer_end_label: str


def read_records() -> tuple[int, list[PageRecord]]:
    reader = PdfReader(str(PDF_PATH))
    records: list[PageRecord] = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        label_match = LABEL_RE.search(text)
        if not label_match:
            continue
        raw_titles = [
            match.group(1)
            for match in STANDALONE_TITLE_RE.finditer(text[:2600])
        ]
        if not raw_titles:
            raw_titles = [match.group(1) for match in TITLE_RE.finditer(text[:2200])]
        titles = tuple(dict.fromkeys(re.sub(r"\s+", "", title) for title in raw_titles))
        is_answer_start = bool(titles and ANSWER_RE.search(text[:1600]))
        records.append(
            PageRecord(
                pdf_page=index,
                chapter=int(label_match.group(1)),
                seq=int(label_match.group(2)),
                titles=titles,
                is_answer_start=is_answer_start,
            )
        )
    return len(reader.pages), records


def label_at_or_before(records: list[PageRecord], pdf_page: int) -> str:
    candidates = [record for record in records if record.pdf_page <= pdf_page]
    if not candidates:
        return ""
    return candidates[-1].printed_label


def iter_chapter_segments(records: list[PageRecord]) -> list[list[PageRecord]]:
    segments: list[list[PageRecord]] = []
    current: list[PageRecord] = []
    previous_chapter: int | None = None
    for record in sorted(records, key=lambda item: item.pdf_page):
        if current and record.chapter != previous_chapter:
            segments.append(current)
            current = []
        current.append(record)
        previous_chapter = record.chapter
    if current:
        segments.append(current)
    return segments


def build_rows(records: list[PageRecord]) -> list[GameRow]:
    rows: list[GameRow] = []
    for segment_index, chapter_records in enumerate(iter_chapter_segments(records), start=1):
        chapter = chapter_records[0].chapter
        answer_starts = [r for r in chapter_records if r.title and r.is_answer_start]
        first_answer_page = answer_starts[0].pdf_page if answer_starts else 10**9
        problem_starts = [
            StartRecord(r.pdf_page, r.printed_label, title)
            for r in chapter_records
            if not r.is_answer_start and r.pdf_page < first_answer_page
            for title in r.titles
        ]
        answer_starts = [
            StartRecord(r.pdf_page, r.printed_label, title)
            for r in answer_starts
            for title in r.titles
        ]

        if len(problem_starts) != len(answer_starts):
            raise RuntimeError(
                f"segment {segment_index} chapter {chapter}: problem starts {len(problem_starts)} != answer starts {len(answer_starts)}"
            )

        last_chapter_labeled_page = chapter_records[-1].pdf_page
        for index, problem in enumerate(problem_starts):
            answer = answer_starts[index]
            next_problem = problem_starts[index + 1] if index + 1 < len(problem_starts) else None
            next_answer = answer_starts[index + 1] if index + 1 < len(answer_starts) else None

            problem_end = max(problem.pdf_page, next_problem.pdf_page - 1) if next_problem else (first_answer_page - 1)
            answer_end = max(answer.pdf_page, next_answer.pdf_page - 1) if next_answer else last_chapter_labeled_page

            if problem_end < problem.pdf_page:
                raise RuntimeError(f"{problem.printed_label}: problem range is inverted")
            if answer_end < answer.pdf_page:
                raise RuntimeError(f"{answer.printed_label}: answer range is inverted")

            rows.append(
                GameRow(
                    problem_id=f"計算{len(rows) + 1:03d}_{problem.printed_label}",
                    title=problem.title,
                    problem_start_page=problem.pdf_page,
                    problem_end_page=problem_end,
                    answer_start_page=answer.pdf_page,
                    answer_end_page=answer_end,
                    problem_start_label=problem.printed_label,
                    problem_end_label=label_at_or_before(chapter_records, problem_end),
                    answer_start_label=answer.printed_label,
                    answer_end_label=label_at_or_before(chapter_records, answer_end),
                )
            )
    return rows


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def write_workbook(rows: list[GameRow], total_pages: int) -> None:
    wb = openpyxl.Workbook()
    ws_q = wb.active
    ws_q.title = "問題"

    question_headers = ["問題番号"]
    for idx in range(1, 21):
        question_headers.extend([f"{idx}問目の回答範囲", f"{idx}問目の正解"])
    ws_q.append(question_headers)

    accepted = "1|○|〇|まる|ok|OK|確認"
    for row in rows:
        data = [row.problem_id, "自己採点", accepted]
        data.extend([""] * (len(question_headers) - len(data)))
        ws_q.append(data)

    ws_p = wb.create_sheet("ページ情報")
    ws_p.append(
        [
            "問題番号",
            "タイトル",
            "PDF内問題開始ページ",
            "PDF内問題終了ページ",
            "PDF内解答開始ページ",
            "PDF内解答終了ページ",
            "書籍印刷問題開始ページ",
            "書籍印刷問題終了ページ",
            "書籍印刷解答開始ページ",
            "書籍印刷解答終了ページ",
        ]
    )
    for row in rows:
        ws_p.append(
            [
                row.problem_id,
                row.title,
                row.problem_start_page,
                row.problem_end_page,
                row.answer_start_page,
                row.answer_end_page,
                row.problem_start_label,
                row.problem_end_label,
                row.answer_start_label,
                row.answer_end_label,
            ]
        )

    ws_h = wb.create_sheet("使い方")
    ws_h.append(["項目", "内容"])
    ws_h.append(["対象PDF", PDF_PATH.name])
    ws_h.append(["総PDFページ数", total_pages])
    ws_h.append(["登録問題数", len(rows)])
    ws_h.append(["採点方式", "計算問題集のため、解答ページを見て自己採点します。正解扱いにする場合は 1 / ○ / 〇 / まる / ok / OK / 確認 のいずれかを入力してください。0 は不正解入力として使えます。"])
    ws_h.append(["ページ範囲", "問題文または解答解説が複数ページにまたがる場合は、終了ページまで含めて表示されるように設定しています。"])

    for ws in [ws_q, ws_p, ws_h]:
        style_header(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = max(len(str(ws.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(1, min(ws.max_row, 200) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)

    ws_q.column_dimensions["A"].width = 16
    ws_p.column_dimensions["A"].width = 16
    ws_p.column_dimensions["B"].width = 34
    ws_h.column_dimensions["A"].width = 18
    ws_h.column_dimensions["B"].width = 90

    wb.save(OUT_PATH)


def main() -> None:
    total_pages, records = read_records()
    rows = build_rows(records)
    write_workbook(rows, total_pages)

    multi_problem = [r for r in rows if r.problem_end_page > r.problem_start_page]
    multi_answer = [r for r in rows if r.answer_end_page > r.answer_start_page]
    print(f"pdf_pages={total_pages}")
    print(f"labeled_pages={len(records)}")
    print(f"game_rows={len(rows)}")
    print(f"multi_page_problem_rows={len(multi_problem)}")
    print(f"multi_page_answer_rows={len(multi_answer)}")
    print(f"output={OUT_PATH}")
    print("sample_multi_problem_rows=")
    for row in multi_problem[:10]:
        print(
            f"  {row.problem_id}: problem {row.problem_start_page}-{row.problem_end_page} "
            f"({row.problem_start_label}-{row.problem_end_label}), "
            f"answer {row.answer_start_page}-{row.answer_end_page} "
            f"({row.answer_start_label}-{row.answer_end_label})"
        )
    print("sample_multi_answer_rows=")
    for row in multi_answer[:10]:
        print(
            f"  {row.problem_id}: problem {row.problem_start_page}-{row.problem_end_page} "
            f"({row.problem_start_label}-{row.problem_end_label}), "
            f"answer {row.answer_start_page}-{row.answer_end_page} "
            f"({row.answer_start_label}-{row.answer_end_label})"
        )


if __name__ == "__main__":
    main()
