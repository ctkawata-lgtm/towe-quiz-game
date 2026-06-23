from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database")
PDF_PATH = BASE / "24_kakomonn_unei.pdf"
OUTPUT_PATH = BASE / "24_kakomonn_unei_pdf_game.xlsx"

YEARS = [
    ("R4", "令和4年度", 159),
    ("R3", "令和3年度", 183),
    ("R2", "令和2年度", 207),
    ("R1", "令和元年度", 229),
    ("H30", "平成30年度", 251),
    ("H29", "平成29年度", 275),
]


def page_text(reader: PdfReader, page_no: int) -> str:
    return (reader.pages[page_no - 1].extract_text() or "").replace("\n", " ")


def normalize_table_text(text: str) -> str:
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"(\d+)\s*－", r" \1 － ", text)
    text = re.sub(r"(\d+)\s*設問", r" \1 設問", text)
    text = re.sub(r"設問\s*(\d+)", r"設問\1 ", text)
    return text


def parse_answer_table(text: str) -> dict[int, list[tuple[str, str]]]:
    text = normalize_table_text(text)
    body = text.split("問題 設問 正解 難易度 配点", 1)[-1].split("合計", 1)[0]
    answers: dict[int, list[tuple[str, str]]] = {}
    pattern = re.compile(r"(\d+)\s+(.+?)(?=\s+\d+\s+(?:－|設問1)|\s*$)")
    for match in pattern.finditer(body):
        num = int(match.group(1))
        chunk = match.group(2).strip()
        sub_answers = re.findall(r"設問\s*(\d+)\s*([アイウエオ])", chunk)
        if sub_answers:
            answers[num] = [(f"設問{sub}", ans) for sub, ans in sub_answers]
            continue
        single = re.search(r"－\s*([アイウエオ－])", chunk)
        if not single:
            continue
        ans = single.group(1)
        answers[num] = [("正解", "なし" if ans == "－" else ans)]
    return answers


def collect_question_starts(reader: PdfReader, prefix: str) -> dict[int, int]:
    starts: dict[int, int] = {}
    marker = f"{prefix}運営 問"
    for page_no in range(1, len(reader.pages) + 1):
        text = page_text(reader, page_no)
        if marker not in text:
            continue
        for match in re.finditer(r"第\s*(\d+)\s*問", text):
            starts.setdefault(int(match.group(1)), page_no)
    return starts


def collect_answer_starts(reader: PdfReader, prefix: str) -> dict[int, int]:
    starts: dict[int, int] = {}
    marker = f"{prefix}運営 解答"
    for page_no in range(1, len(reader.pages) + 1):
        text = page_text(reader, page_no)
        if marker not in text:
            continue
        for match in re.finditer(r"第\s*(\d+)\s*問", text):
            starts.setdefault(int(match.group(1)), page_no)
    return starts


def page_end(starts: dict[int, int], num: int) -> int:
    current = starts[num]
    next_pages = [page for q, page in starts.items() if q > num]
    if not next_pages:
        return current
    next_page = min(next_pages)
    return current if next_page == current else next_page - 1


def build_workbook() -> None:
    reader = PdfReader(str(PDF_PATH))
    wb = Workbook()
    ws_answers = wb.active
    ws_answers.title = "問題"
    ws_pages = wb.create_sheet("ページ情報")
    ws_usage = wb.create_sheet("使い方")

    answer_headers = ["問題番号"]
    for i in range(1, 21):
        answer_headers += [f"{i}問目の回答範囲", f"{i}問目の正解"]
    page_headers = [
        "問題番号",
        "PDF内問題開始ページ",
        "PDF内問題終了ページ",
        "PDF内解答開始ページ",
        "PDF内解答終了ページ",
        "書籍印刷問題開始ページ",
        "書籍印刷問題終了ページ",
        "書籍印刷解答開始ページ",
        "書籍印刷解答終了ページ",
    ]
    ws_answers.append(answer_headers)
    ws_pages.append(page_headers)

    expected_total = 0
    for prefix, _label, table_page in YEARS:
        correct = parse_answer_table(page_text(reader, table_page))
        question_starts = collect_question_starts(reader, prefix)
        answer_starts = collect_answer_starts(reader, prefix)
        expected_total += len(question_starts)

        missing_answers = sorted(set(question_starts) - set(correct))
        missing_pages = sorted(set(question_starts) - set(answer_starts))
        if missing_answers:
            raise RuntimeError(f"{prefix}: missing correct answers {missing_answers}")
        if missing_pages:
            raise RuntimeError(f"{prefix}: missing answer pages {missing_pages}")

        for num in sorted(question_starts):
            qid = f"{prefix}-{num:02d}"
            row = [qid] + [None] * 40
            for idx, (scope, ans) in enumerate(correct[num], start=1):
                row[1 + (idx - 1) * 2] = scope
                row[2 + (idx - 1) * 2] = ans
            ws_answers.append(row)

            q_start = question_starts[num]
            q_end = page_end(question_starts, num)
            a_start = answer_starts[num]
            a_end = page_end(answer_starts, num)
            ws_pages.append([
                qid,
                q_start,
                q_end,
                a_start,
                a_end,
                q_start - 4,
                q_end - 4,
                a_start - 4,
                a_end - 4,
            ])

    ws_usage.append(["24_kakomonn_unei.pdf PDFゲーム用データ", "PDF + Excel モード"])
    ws_usage.append(["対象PDF", PDF_PATH.name])
    ws_usage.append(["問題数", expected_total])
    ws_usage.append(["問題シート", "問題番号、回答範囲、正解を収録。小問は同じ行に設問1、設問2として収録。"])
    ws_usage.append(["ページ情報シート", "PDF内ページと書籍印刷ページを対応付け。問題・解答とも見開き表示に対応。"])
    ws_usage.append(["注意", "図表問題もページ単位で収録。ゲーム側ではPDF本文を見ながら解答してください。"])

    style_workbook(wb)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)
    print(f"questions={expected_total}")


def style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            width = 14
            if col == 1:
                width = 13
            elif ws.title == "問題":
                width = 18 if col % 2 == 0 else 12
            elif ws.title == "ページ情報":
                width = 20
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.auto_filter.ref = ws.dimensions


if __name__ == "__main__":
    build_workbook()
