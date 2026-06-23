from pathlib import Path

import openpyxl


PATH = Path(r"C:\Users\kawat\Documents\Codex\2026-05-29\quiz_database\26財務会計論（簿記）cpa、過去問_pdf_game.xlsx")
PDF_TOTAL_PAGES = 1299


def main() -> None:
    wb = openpyxl.load_workbook(PATH, data_only=True)
    ws_q = wb["問題"]
    ws_p = wb["ページ情報"]
    ids = set()
    bad = []
    too_many = []
    long_answers = []
    for row in ws_q.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        problem_id = row[0]
        ids.add(problem_id)
        count = 0
        for index in range(1, 21):
            answer = row[index * 2] if index * 2 < len(row) else None
            if answer:
                count += 1
                if len(str(answer)) > 16:
                    long_answers.append((problem_id, index, answer))
        if count > 20:
            too_many.append((problem_id, count))

    page_ids = set()
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        problem_id = row[0]
        page_ids.add(problem_id)
        q_start, q_end, a_start, a_end = row[2], row[3], row[4], row[5]
        if not (1 <= q_start <= q_end <= PDF_TOTAL_PAGES and 1 <= a_start <= a_end <= PDF_TOTAL_PAGES):
            bad.append((problem_id, q_start, q_end, a_start, a_end))
        if a_start <= q_end:
            bad.append((problem_id, q_start, q_end, a_start, a_end, "overlap"))

    skipped = wb["要確認"].max_row - 1
    print(f"sheets={wb.sheetnames}")
    print(f"question_rows={ws_q.max_row - 1}")
    print(f"page_rows={ws_p.max_row - 1}")
    print(f"id_mismatch={len(ids ^ page_ids)}")
    print(f"bad_ranges={len(bad)}")
    print(f"too_many={len(too_many)}")
    print(f"long_answers={len(long_answers)}")
    print(f"confirm_rows={skipped}")
    print(f"first_question={[cell.value for cell in ws_q[2]][:9]}")
    print(f"first_page_info={[cell.value for cell in ws_p[2]][:6]}")
    if ids ^ page_ids or bad or too_many or long_answers:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
